"""
Cross-source listing store updater
=================================

What it does:
- finds the latest top-level scraper JSON outputs
- ingests them into a SQLite store
- tracks new / updated / unchanged / removed listings per source
- exports a consolidated workbook for active and changed listings

Run:
  python3 listing_store/update_store.py
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sqlite3
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRAPERS_DIR = PROJECT_ROOT / "scrapers"
OUTPUT_DIR = PROJECT_ROOT / "listing_store" / "outputs"
DB_PATH = OUTPUT_DIR / "listings.sqlite3"
WORKBOOK_PATH = OUTPUT_DIR / "lyon_master_listings.xlsx"
SUMMARY_JSON_PATH = OUTPUT_DIR / "latest_pipeline_summary.json"
ACTIVE_JSON_PATH = OUTPUT_DIR / "active_listings.json"
ACTIVE_CSV_PATH = OUTPUT_DIR / "active_listings.csv"
NEW_JSON_PATH = OUTPUT_DIR / "new_in_run.json"
UPDATED_JSON_PATH = OUTPUT_DIR / "updated_in_run.json"
REMOVED_JSON_PATH = OUTPUT_DIR / "removed_in_run.json"


EXPECTED_OUTPUTS = {
    "immojeune": SCRAPERS_DIR / "immojeune" / "outputs" / "immojeune_lyon_shared_housing.json",
    "la_carte_des_colocs": SCRAPERS_DIR / "la_carte_des_colocs" / "outputs" / "la_carte_des_colocs_lyon.json",
    "location_etudiant": SCRAPERS_DIR / "location_etudiant" / "outputs" / "location_etudiant_lyon_shared_residences.json",
    "studapart": SCRAPERS_DIR / "studapart" / "outputs" / "studapart_lyon_shared_listings.json",
}


@dataclass
class PayloadBundle:
    source: str
    path: Path
    payload: dict[str, Any]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def stable_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def fingerprint_for(listing: dict[str, Any], normalized: dict[str, Any]) -> str:
    listing_copy = dict(listing)
    normalized_copy = dict(normalized)
    listing_copy.pop("scraped_at", None)
    normalized_copy.pop("scraped_at", None)
    blob = stable_json({"listing": listing_copy, "normalized": normalized_copy})
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def discover_payloads() -> list[PayloadBundle]:
    bundles: list[PayloadBundle] = []
    for source, path in EXPECTED_OUTPUTS.items():
        if not path.exists():
            continue
        payload = json.loads(path.read_text(encoding="utf-8"))
        bundles.append(PayloadBundle(source=source, path=path, payload=payload))
    bundles.sort(key=lambda item: item.source)
    return bundles


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA foreign_keys = ON;

        CREATE TABLE IF NOT EXISTS pipeline_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            started_at TEXT NOT NULL,
            finished_at TEXT,
            source_count INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS source_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pipeline_run_id INTEGER NOT NULL,
            source TEXT NOT NULL,
            input_path TEXT NOT NULL,
            started_at TEXT NOT NULL,
            listing_count INTEGER NOT NULL DEFAULT 0,
            meta_json TEXT NOT NULL,
            FOREIGN KEY (pipeline_run_id) REFERENCES pipeline_runs(id)
        );

        CREATE TABLE IF NOT EXISTS listings (
            listing_uid TEXT PRIMARY KEY,
            source TEXT NOT NULL,
            source_listing_id TEXT,
            url TEXT NOT NULL,
            title TEXT NOT NULL,
            price_eur INTEGER,
            postcode TEXT,
            latitude REAL,
            longitude REAL,
            first_seen_at TEXT NOT NULL,
            last_seen_at TEXT NOT NULL,
            removed_at TEXT,
            is_active INTEGER NOT NULL DEFAULT 1,
            fingerprint TEXT NOT NULL,
            latest_status TEXT NOT NULL,
            latest_raw_json TEXT NOT NULL,
            latest_normalized_json TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS listing_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pipeline_run_id INTEGER NOT NULL,
            source_run_id INTEGER NOT NULL,
            listing_uid TEXT NOT NULL,
            source TEXT NOT NULL,
            status TEXT NOT NULL,
            seen_at TEXT NOT NULL,
            fingerprint TEXT NOT NULL,
            raw_json TEXT NOT NULL,
            normalized_json TEXT NOT NULL,
            FOREIGN KEY (pipeline_run_id) REFERENCES pipeline_runs(id),
            FOREIGN KEY (source_run_id) REFERENCES source_runs(id)
        );
        """
    )
    conn.commit()


def listing_uid(source: str, listing: dict[str, Any], normalized: dict[str, Any]) -> str:
    source_listing_id = listing.get("source_listing_id")
    if source_listing_id:
        return f"{source}::{source_listing_id}"
    url = normalized.get("url") or listing.get("url")
    if url:
        return f"{source}::{url}"
    title = normalized.get("title") or listing.get("title") or "untitled"
    return f"{source}::{title}"


def source_listing_id_value(listing: dict[str, Any]) -> str | None:
    value = listing.get("source_listing_id")
    return str(value) if value is not None else None


def choose_title(listing: dict[str, Any], normalized: dict[str, Any]) -> str:
    return str(normalized.get("title") or listing.get("title") or "Untitled listing")


def choose_url(listing: dict[str, Any], normalized: dict[str, Any]) -> str:
    return str(normalized.get("url") or listing.get("url") or "")


def choose_price(normalized: dict[str, Any]) -> int | None:
    value = normalized.get("price_eur")
    return int(value) if isinstance(value, int) else None


def choose_postcode(normalized: dict[str, Any], listing: dict[str, Any]) -> str | None:
    value = normalized.get("postcode") or listing.get("postcode")
    return str(value) if value not in (None, "") else None


def choose_float(value: Any) -> float | None:
    return float(value) if isinstance(value, (int, float)) else None


def pair_records(payload: dict[str, Any]) -> list[tuple[dict[str, Any], dict[str, Any]]]:
    listings = payload.get("listings") or []
    normalized = payload.get("normalized_listings") or []
    pairs: list[tuple[dict[str, Any], dict[str, Any]]] = []
    max_len = max(len(listings), len(normalized))
    for index in range(max_len):
        listing = listings[index] if index < len(listings) and isinstance(listings[index], dict) else {}
        norm = normalized[index] if index < len(normalized) and isinstance(normalized[index], dict) else {}
        pairs.append((listing, norm))
    return pairs


def ingest_source_run(
    conn: sqlite3.Connection,
    pipeline_run_id: int,
    bundle: PayloadBundle,
    seen_at: str,
) -> dict[str, int]:
    pairs = pair_records(bundle.payload)
    meta_json = stable_json(bundle.payload.get("meta") or {})
    cursor = conn.execute(
        """
        INSERT INTO source_runs (pipeline_run_id, source, input_path, started_at, listing_count, meta_json)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (pipeline_run_id, bundle.source, str(bundle.path), seen_at, len(pairs), meta_json),
    )
    source_run_id = int(cursor.lastrowid)

    current_uids: set[str] = set()
    status_counts: Counter[str] = Counter()

    for listing, normalized in pairs:
        uid = listing_uid(bundle.source, listing, normalized)
        current_uids.add(uid)
        fingerprint = fingerprint_for(listing, normalized)
        existing = conn.execute(
            """
            SELECT fingerprint, first_seen_at, latest_raw_json, latest_normalized_json
            FROM listings
            WHERE listing_uid = ?
            """,
            (uid,),
        ).fetchone()

        if existing is None:
            status = "new"
            first_seen_at = seen_at
        else:
            first_seen_at = str(existing[1])
            status = "updated" if str(existing[0]) != fingerprint else "unchanged"

        raw_json = stable_json(listing)
        normalized_json = stable_json(normalized)
        conn.execute(
            """
            INSERT INTO listings (
                listing_uid, source, source_listing_id, url, title, price_eur, postcode,
                latitude, longitude, first_seen_at, last_seen_at, removed_at, is_active,
                fingerprint, latest_status, latest_raw_json, latest_normalized_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, 1, ?, ?, ?, ?)
            ON CONFLICT(listing_uid) DO UPDATE SET
                source = excluded.source,
                source_listing_id = excluded.source_listing_id,
                url = excluded.url,
                title = excluded.title,
                price_eur = excluded.price_eur,
                postcode = excluded.postcode,
                latitude = excluded.latitude,
                longitude = excluded.longitude,
                last_seen_at = excluded.last_seen_at,
                removed_at = NULL,
                is_active = 1,
                fingerprint = excluded.fingerprint,
                latest_status = excluded.latest_status,
                latest_raw_json = excluded.latest_raw_json,
                latest_normalized_json = excluded.latest_normalized_json
            """,
            (
                uid,
                bundle.source,
                source_listing_id_value(listing),
                choose_url(listing, normalized),
                choose_title(listing, normalized),
                choose_price(normalized),
                choose_postcode(normalized, listing),
                choose_float(normalized.get("latitude")),
                choose_float(normalized.get("longitude")),
                first_seen_at,
                seen_at,
                fingerprint,
                status,
                raw_json,
                normalized_json,
            ),
        )
        conn.execute(
            """
            INSERT INTO listing_history (
                pipeline_run_id, source_run_id, listing_uid, source, status,
                seen_at, fingerprint, raw_json, normalized_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                pipeline_run_id,
                source_run_id,
                uid,
                bundle.source,
                status,
                seen_at,
                fingerprint,
                raw_json,
                normalized_json,
            ),
        )
        status_counts[status] += 1

    previously_active = conn.execute(
        "SELECT listing_uid, fingerprint, latest_raw_json, latest_normalized_json FROM listings WHERE source = ? AND is_active = 1",
        (bundle.source,),
    ).fetchall()
    for row in previously_active:
        uid = str(row[0])
        if uid in current_uids:
            continue
        conn.execute(
            """
            UPDATE listings
            SET is_active = 0, removed_at = ?, latest_status = ?
            WHERE listing_uid = ?
            """,
            (seen_at, "removed", uid),
        )
        conn.execute(
            """
            INSERT INTO listing_history (
                pipeline_run_id, source_run_id, listing_uid, source, status,
                seen_at, fingerprint, raw_json, normalized_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                pipeline_run_id,
                source_run_id,
                uid,
                bundle.source,
                "removed",
                seen_at,
                str(row[1]),
                str(row[2]),
                str(row[3]),
            ),
        )
        status_counts["removed"] += 1

    conn.commit()
    return dict(status_counts)


def human_address(source: str, raw: dict[str, Any]) -> str:
    for key in ("full_address", "address", "address_text", "street", "location_text", "city"):
        value = raw.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def human_city(source: str, raw: dict[str, Any]) -> str:
    value = raw.get("city")
    if isinstance(value, str) and value.strip():
        return value.strip()
    if source == "immojeune":
        postcode = raw.get("postcode")
        if isinstance(postcode, str) and postcode.startswith("69"):
            return "Lyon area"
    return ""


def human_availability(raw: dict[str, Any]) -> str:
    for key in ("availability", "available_from"):
        value = raw.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def human_extra_summary(source: str, raw: dict[str, Any]) -> str:
    if source == "immojeune":
        parts = [raw.get("student_label"), raw.get("mixed_household"), raw.get("nearby_schools")]
    elif source == "la_carte_des_colocs":
        parts = [raw.get("listing_type"), raw.get("lodging_type_label"), raw.get("current_housemates")]
    elif source == "location_etudiant":
        unit_types = raw.get("unit_types")
        parts = [raw.get("operator"), ", ".join(unit_types) if isinstance(unit_types, list) else unit_types]
    elif source == "studapart":
        rented_by_room = raw.get("rented_by_room")
        rented_label = "rented by room" if rented_by_room is True else ""
        parts = [raw.get("announcement_type"), raw.get("property_type"), raw.get("max_tenants"), rented_label]
    else:
        parts = []
    clean_parts = [str(part) for part in parts if part not in (None, "", "Not listed", "Unknown")]
    return " | ".join(clean_parts)


def active_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT
            listing_uid, source, source_listing_id, url, title, price_eur, postcode,
            latitude, longitude, first_seen_at, last_seen_at, latest_status,
            latest_raw_json, latest_normalized_json
        FROM listings
        WHERE is_active = 1
        ORDER BY COALESCE(price_eur, 999999), source, title
        """
    ).fetchall()
    result: list[dict[str, Any]] = []
    for row in rows:
        raw = json.loads(str(row[12]))
        normalized = json.loads(str(row[13]))
        result.append(
            {
                "listing_uid": row[0],
                "source": row[1],
                "source_listing_id": row[2],
                "url": row[3],
                "title": row[4],
                "price_eur": row[5],
                "postcode": row[6],
                "latitude": row[7],
                "longitude": row[8],
                "first_seen_at": row[9],
                "last_seen_at": row[10],
                "latest_status": row[11],
                "city": human_city(str(row[1]), raw),
                "address": human_address(str(row[1]), raw),
                "availability": human_availability(raw),
                "extra_summary": human_extra_summary(str(row[1]), raw),
                "raw": raw,
                "normalized": normalized,
            }
        )
    return result


def changed_rows(conn: sqlite3.Connection, pipeline_run_id: int, statuses: tuple[str, ...]) -> list[dict[str, Any]]:
    placeholders = ", ".join("?" for _ in statuses)
    rows = conn.execute(
        f"""
        SELECT listing_uid, source, status, seen_at, raw_json, normalized_json
        FROM listing_history
        WHERE pipeline_run_id = ? AND status IN ({placeholders})
        ORDER BY source, status, listing_uid
        """,
        (pipeline_run_id, *statuses),
    ).fetchall()
    result: list[dict[str, Any]] = []
    for row in rows:
        raw = json.loads(str(row[4]))
        normalized = json.loads(str(row[5]))
        result.append(
            {
                "listing_uid": row[0],
                "source": row[1],
                "status": row[2],
                "seen_at": row[3],
                "title": normalized.get("title") or raw.get("title") or "",
                "url": normalized.get("url") or raw.get("url") or "",
                "price_eur": normalized.get("price_eur"),
                "postcode": normalized.get("postcode") or raw.get("postcode"),
                "address": human_address(str(row[1]), raw),
                "availability": human_availability(raw),
                "extra_summary": human_extra_summary(str(row[1]), raw),
            }
        )
    return result


def source_run_rows(conn: sqlite3.Connection, pipeline_run_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT source, input_path, started_at, listing_count, meta_json
        FROM source_runs
        WHERE pipeline_run_id = ?
        ORDER BY source
        """,
        (pipeline_run_id,),
    ).fetchall()
    result = []
    for row in rows:
        result.append(
            {
                "source": row[0],
                "input_path": row[1],
                "started_at": row[2],
                "listing_count": row[3],
                "meta_json": row[4],
            }
        )
    return result


def write_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_workbook(
    path: Path,
    active: list[dict[str, Any]],
    new_rows: list[dict[str, Any]],
    updated_rows: list[dict[str, Any]],
    removed_rows: list[dict[str, Any]],
    run_rows: list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    active_ws = wb.active
    active_ws.title = "Active Listings"
    active_headers = [
        "Source",
        "Listing UID",
        "Source Listing ID",
        "Title",
        "Price EUR",
        "Postcode",
        "City",
        "Address",
        "Availability",
        "Latitude",
        "Longitude",
        "First Seen",
        "Last Seen",
        "Latest Status",
        "Extra Summary",
        "URL",
    ]
    active_rows_data = [
        [
            row["source"],
            row["listing_uid"],
            row["source_listing_id"],
            row["title"],
            row["price_eur"],
            row["postcode"],
            row["city"],
            row["address"],
            row["availability"],
            row["latitude"],
            row["longitude"],
            row["first_seen_at"],
            row["last_seen_at"],
            row["latest_status"],
            row["extra_summary"],
            row["url"],
        ]
        for row in active
    ]
    write_sheet(active_ws, active_headers, active_rows_data)

    for ws_name, rows in {
        "New In Run": new_rows,
        "Updated In Run": updated_rows,
        "Removed In Run": removed_rows,
    }.items():
        ws = wb.create_sheet(ws_name)
        headers = ["Source", "Status", "Listing UID", "Title", "Price EUR", "Postcode", "Address", "Availability", "Extra Summary", "URL", "Seen At"]
        data_rows = [
            [
                row["source"],
                row["status"],
                row["listing_uid"],
                row["title"],
                row["price_eur"],
                row["postcode"],
                row["address"],
                row["availability"],
                row["extra_summary"],
                row["url"],
                row["seen_at"],
            ]
            for row in rows
        ]
        write_sheet(ws, headers, data_rows)

    runs_ws = wb.create_sheet("Source Runs")
    run_headers = ["Source", "Input Path", "Started At", "Listing Count", "Meta JSON"]
    run_data = [[row["source"], row["input_path"], row["started_at"], row["listing_count"], row["meta_json"]] for row in run_rows]
    write_sheet(runs_ws, run_headers, run_data)

    for ws in wb.worksheets:
        widths = {
            "A": 18,
            "B": 44,
            "C": 24,
            "D": 42,
            "E": 12,
            "F": 12,
            "G": 22,
            "H": 34,
            "I": 18,
            "J": 12,
            "K": 12,
            "L": 28,
            "M": 28,
            "N": 16,
            "O": 48,
            "P": 58,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    wb.save(path)


def export_active_json(path: Path, active: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [
        {
            "source": row["source"],
            "listing_uid": row["listing_uid"],
            "source_listing_id": row["source_listing_id"],
            "title": row["title"],
            "price_eur": row["price_eur"],
            "postcode": row["postcode"],
            "city": row["city"],
            "address": row["address"],
            "availability": row["availability"],
            "latitude": row["latitude"],
            "longitude": row["longitude"],
            "first_seen_at": row["first_seen_at"],
            "last_seen_at": row["last_seen_at"],
            "latest_status": row["latest_status"],
            "extra_summary": row["extra_summary"],
            "url": row["url"],
        }
        for row in active
    ]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_active_csv(path: Path, active: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "source",
                "listing_uid",
                "source_listing_id",
                "title",
                "price_eur",
                "postcode",
                "city",
                "address",
                "availability",
                "latitude",
                "longitude",
                "first_seen_at",
                "last_seen_at",
                "latest_status",
                "extra_summary",
                "url",
            ]
        )
        for row in active:
            writer.writerow(
                [
                    row["source"],
                    row["listing_uid"],
                    row["source_listing_id"],
                    row["title"],
                    row["price_eur"],
                    row["postcode"],
                    row["city"],
                    row["address"],
                    row["availability"],
                    row["latitude"],
                    row["longitude"],
                    row["first_seen_at"],
                    row["last_seen_at"],
                    row["latest_status"],
                    row["extra_summary"],
                    row["url"],
                ]
            )


def export_rows_json(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def build_summary(
    pipeline_run_id: int,
    bundles: list[PayloadBundle],
    status_counts_by_source: dict[str, dict[str, int]],
    active_count: int,
    new_count: int,
    updated_count: int,
    removed_count: int,
) -> dict[str, Any]:
    return {
        "pipeline_run_id": pipeline_run_id,
        "generated_at": utc_now_iso(),
        "sources_ingested": [bundle.source for bundle in bundles],
        "status_counts_by_source": status_counts_by_source,
        "active_count": active_count,
        "new_count": new_count,
        "updated_count": updated_count,
        "removed_count": removed_count,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Update the cross-source listing store.")
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH))
    parser.add_argument("--summary-json", default=str(SUMMARY_JSON_PATH))
    parser.add_argument("--active-json", default=str(ACTIVE_JSON_PATH))
    parser.add_argument("--active-csv", default=str(ACTIVE_CSV_PATH))
    parser.add_argument("--new-json", default=str(NEW_JSON_PATH))
    parser.add_argument("--updated-json", default=str(UPDATED_JSON_PATH))
    parser.add_argument("--removed-json", default=str(REMOVED_JSON_PATH))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    bundles = discover_payloads()
    if not bundles:
        raise SystemExit("No scraper JSON payloads found to ingest.")

    db_path = Path(args.db).expanduser().resolve()
    workbook_path = Path(args.workbook).expanduser().resolve()
    summary_json_path = Path(args.summary_json).expanduser().resolve()
    active_json_path = Path(args.active_json).expanduser().resolve()
    active_csv_path = Path(args.active_csv).expanduser().resolve()
    new_json_path = Path(args.new_json).expanduser().resolve()
    updated_json_path = Path(args.updated_json).expanduser().resolve()
    removed_json_path = Path(args.removed_json).expanduser().resolve()

    conn = sqlite3.connect(db_path)
    try:
        ensure_schema(conn)
        started_at = utc_now_iso()
        cursor = conn.execute(
            "INSERT INTO pipeline_runs (started_at, source_count) VALUES (?, ?)",
            (started_at, len(bundles)),
        )
        pipeline_run_id = int(cursor.lastrowid)

        status_counts_by_source: dict[str, dict[str, int]] = {}
        for bundle in bundles:
            status_counts_by_source[bundle.source] = ingest_source_run(
                conn=conn,
                pipeline_run_id=pipeline_run_id,
                bundle=bundle,
                seen_at=started_at,
            )

        finished_at = utc_now_iso()
        conn.execute(
            "UPDATE pipeline_runs SET finished_at = ? WHERE id = ?",
            (finished_at, pipeline_run_id),
        )
        conn.commit()

        active = active_rows(conn)
        new_rows = changed_rows(conn, pipeline_run_id, ("new",))
        updated_rows = changed_rows(conn, pipeline_run_id, ("updated",))
        removed_rows = changed_rows(conn, pipeline_run_id, ("removed",))
        run_rows = source_run_rows(conn, pipeline_run_id)

        export_workbook(
            workbook_path,
            active=active,
            new_rows=new_rows,
            updated_rows=updated_rows,
            removed_rows=removed_rows,
            run_rows=run_rows,
        )

        summary = build_summary(
            pipeline_run_id=pipeline_run_id,
            bundles=bundles,
            status_counts_by_source=status_counts_by_source,
            active_count=len(active),
            new_count=len(new_rows),
            updated_count=len(updated_rows),
            removed_count=len(removed_rows),
        )
        summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        export_active_json(active_json_path, active)
        export_active_csv(active_csv_path, active)
        export_rows_json(new_json_path, new_rows)
        export_rows_json(updated_json_path, updated_rows)
        export_rows_json(removed_json_path, removed_rows)
    finally:
        conn.close()

    print(f"Updated SQLite store: {db_path}")
    print(f"Saved consolidated workbook: {workbook_path}")
    print(f"Saved pipeline summary: {summary_json_path}")
    print(f"Saved active listings JSON: {active_json_path}")
    print(f"Saved active listings CSV: {active_csv_path}")
    print(f"Saved new listings JSON: {new_json_path}")
    print(f"Saved updated listings JSON: {updated_json_path}")
    print(f"Saved removed listings JSON: {removed_json_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
