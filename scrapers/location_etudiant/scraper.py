"""
Location Étudiant Lyon shared-friendly residence scraper
=======================================================

What it does:
- fetches the Lyon residences page
- keeps only residences with meaningful shared-living signals
- enriches strong candidates from their detail pages
- exports Excel + JSON

Run:
  python3 scrapers/location_etudiant/scraper.py
  python3 scrapers/location_etudiant/scraper.py --max-results 10

Dependencies:
  pip install beautifulsoup4 openpyxl
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from shared.listing_schema import NormalizedListing
from scrapers.location_etudiant.recon import (  # type: ignore[import-not-found]
    DEFAULT_URL,
    RAW_HTML_DIR,
    extract_unit_types,
    fetch_html,
    parse_detail_summary,
    parse_residence_cards,
    read_html,
    shared_label,
    shared_score_for_text,
    utc_now_iso,
)


OUTPUT_DIR = BASE_DIR / "outputs"
DEFAULT_OUTPUT = OUTPUT_DIR / "location_etudiant_lyon_shared_residences.xlsx"
DEFAULT_JSON_OUTPUT = OUTPUT_DIR / "location_etudiant_lyon_shared_residences.json"
MIN_SHARED_SCORE = 2


@dataclass
class Listing:
    source: str
    source_listing_id: str
    url: str
    title: str
    operator: str
    price_eur: int | None
    price_text: str
    address: str
    availability: str
    unit_types: list[str]
    shared_score: int
    shared_label: str
    shared_signals: list[str]
    summary_description: str
    detail_description: str
    scraped_at: str

    def explicit_dict(self) -> dict:
        return {
            "source": self.source,
            "source_listing_id": self.source_listing_id,
            "url": self.url,
            "title": self.title,
            "operator": self.operator,
            "price_eur": self.price_eur,
            "price_text": self.price_text,
            "address": self.address,
            "availability": self.availability,
            "unit_types": self.unit_types,
            "summary_description": self.summary_description,
            "detail_description": self.detail_description,
            "scraped_at": self.scraped_at,
        }

    def normalized(self) -> NormalizedListing:
        return NormalizedListing(
            source=self.source,
            url=self.url,
            title=self.title,
            price_eur=self.price_eur,
            postcode=None,
            latitude=None,
            longitude=None,
            student_occupants="",
            worker_occupants="",
            scraped_at=self.scraped_at,
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape shared-friendly Location Étudiant residences in Lyon.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--json-output", default=str(DEFAULT_JSON_OUTPUT))
    parser.add_argument("--max-results", type=int, default=0)
    parser.add_argument("--min-shared-score", type=int, default=MIN_SHARED_SCORE)
    parser.add_argument("--offline-html", type=Path, help="Use a saved main page instead of fetching live.")
    return parser.parse_args()


def listing_id_from_url(url: str) -> str:
    return url.rstrip("/").split("/")[-1].replace(".html", "")


def build_listing(card: object, detail_description: str, scraped_at: str) -> Listing:
    score, signals = shared_score_for_text(f"{card.title} {detail_description}")
    return Listing(
        source="location_etudiant",
        source_listing_id=listing_id_from_url(card.url),
        url=card.url,
        title=card.title,
        operator=card.operator,
        price_eur=card.price_eur,
        price_text=card.price_text,
        address=card.address,
        availability=card.availability,
        unit_types=extract_unit_types(f"{card.title} {detail_description}"),
        shared_score=score,
        shared_label=shared_label(score, detail_description),
        shared_signals=signals,
        summary_description=card.description,
        detail_description=detail_description,
        scraped_at=scraped_at,
    )


def collect_listings(main_html: str, scraped_at: str, min_shared_score: int) -> tuple[list[Listing], dict]:
    cards = parse_residence_cards(main_html)
    shortlisted = [card for card in cards if card.shared_score >= min_shared_score]
    listings: list[Listing] = []
    detail_cache: dict[str, str] = {}

    for card in shortlisted:
        try:
            detail_html = fetch_html(card.url)
            detail_cache[card.url] = detail_html
            detail = parse_detail_summary(card.url, detail_html)
            detail_description = detail.description or card.description
        except Exception:
            detail_description = card.description

        listing = build_listing(card, detail_description=detail_description, scraped_at=scraped_at)
        if listing.shared_score >= min_shared_score:
            listings.append(listing)

    meta = {
        "card_count": len(cards),
        "shortlisted_from_page": len(shortlisted),
        "kept_count": len(listings),
        "detail_fetch_count": len(detail_cache),
    }
    return listings, meta


def export_excel(path: Path, listings: list[Listing], meta: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shared Residences"

    headers = [
        "Source",
        "Listing ID",
        "Title",
        "URL",
        "Operator",
        "Price EUR",
        "Price Text",
        "Address",
        "Availability",
        "Unit Types",
        "Summary Description",
        "Detail Description",
        "Scraped At",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for listing in listings:
        ws.append(
            [
                listing.source,
                listing.source_listing_id,
                listing.title,
                listing.url,
                listing.operator,
                listing.price_eur,
                listing.price_text,
                listing.address,
                listing.availability,
                ", ".join(listing.unit_types),
                listing.summary_description,
                listing.detail_description,
                listing.scraped_at,
            ]
        )

    widths = {
        "A": 18,
        "B": 18,
        "C": 30,
        "D": 46,
        "E": 24,
        "F": 10,
        "G": 18,
        "H": 34,
        "I": 18,
        "J": 18,
        "K": 80,
        "L": 100,
        "M": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    meta_sheet = wb.create_sheet("Run Meta")
    meta_sheet.append(["Key", "Value"])
    for key, value in meta.items():
        meta_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])
    meta_sheet.column_dimensions["A"].width = 24
    meta_sheet.column_dimensions["B"].width = 48

    wb.save(path)


def main() -> int:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_HTML_DIR.mkdir(parents=True, exist_ok=True)

    if args.offline_html:
        main_html = read_html(args.offline_html)
    else:
        main_html = fetch_html(DEFAULT_URL)
        (RAW_HTML_DIR / "lyon_residences_live_latest.html").write_text(main_html, encoding="utf-8")

    scraped_at = utc_now_iso()
    listings, meta = collect_listings(main_html=main_html, scraped_at=scraped_at, min_shared_score=args.min_shared_score)
    listings.sort(key=lambda item: (-item.shared_score, item.price_eur or 10**9, item.title))

    if args.max_results > 0:
        listings = listings[: args.max_results]

    output_path = Path(args.output).expanduser().resolve()
    json_output_path = Path(args.json_output).expanduser().resolve()

    payload = {
        "source": "location_etudiant",
        "start_url": DEFAULT_URL,
        "meta": meta,
        "listings": [item.explicit_dict() for item in listings],
        "normalized_listings": [item.normalized().to_dict() for item in listings],
    }
    export_excel(output_path, listings, meta)
    json_output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Saved Excel output to: {output_path}")
    print(f"Saved JSON output to: {json_output_path}")
    print(f"Kept {len(listings)} shared-friendly residences from {meta['card_count']} Lyon cards.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
