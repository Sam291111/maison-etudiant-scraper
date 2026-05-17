"""
La Carte des Colocs Lyon scraper
================================

What it does:
- loads the live Lyon search page in a browser session
- captures the site's own listing JSON response
- keeps shared-accommodation focused results
- exports Excel + JSON + raw capture files

Run:
  python3 scrapers/la_carte_des_colocs/scraper.py
  python3 scrapers/la_carte_des_colocs/scraper.py --max-results 20
  python3 scrapers/la_carte_des_colocs/scraper.py --max-batches 2

Dependencies:
  pip install openpyxl playwright
  python3 -m playwright install chromium
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus, urljoin
from urllib.request import Request, urlopen

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from playwright.sync_api import Response, sync_playwright
except ImportError:
    Response = Any  # type: ignore[assignment]
    sync_playwright = None


BASE_URL = "https://www.lacartedescolocs.fr"
START_URL = f"{BASE_URL}/logements/fr/auvergne-rhone-alpes/lyon"
BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from shared.listing_schema import NormalizedListing

OUTPUT_DIR = BASE_DIR / "outputs"
DEFAULT_OUTPUT = OUTPUT_DIR / "la_carte_des_colocs_lyon.xlsx"
DEFAULT_JSON_OUTPUT = OUTPUT_DIR / "la_carte_des_colocs_lyon.json"
RAW_CAPTURE_DIR = OUTPUT_DIR / "raw_capture"
WAIT_MS = 8_000
GEOCODER_URL = "https://data.geopf.fr/geocodage/search"
GEOCODER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}

SHARED_TYPES = {"flatshare", "coliving", "homestay", "student_room", "student_residence", "sublet"}
STUDENT_PATTERNS = [
    r"\betudiant(?:e|es|s)?\b",
    r"\bétudiant(?:e|es|s)?\b",
    r"\bstudent(?:s)?\b",
    r"\bjeune actif\b",
    r"\bjeunes actifs\b",
    r"\bapl\b",
    r"\bcampus\b",
    r"\buniversit",
    r"\becole\b",
    r"\bécole\b",
]
FEMALE_ONLY_PATTERNS = [
    r"\bfille[s]?\s+uniquement\b",
    r"\bfemmes?\s+uniquement\b",
    r"\bentre filles\b",
    r"\bétudiantes?\b",
]
MALE_ONLY_PATTERNS = [
    r"\bgar[cç]ons?\s+uniquement\b",
    r"\bhommes?\s+uniquement\b",
    r"\bentre gars\b",
]


@dataclass
class Listing:
    source: str
    source_listing_id: str
    url: str
    title: str
    price_eur: int | None
    listing_type: str
    lodging_type: str
    lodging_type_label: str
    surface_sqm: int | None
    room_count_label: str
    current_housemates: int | None
    furnished: bool | None
    availability: str
    published_at: str
    published_label: str
    city: str
    street: str
    postcode: str | None
    postcode_source: str
    company_name: str
    description: str
    female_only: str
    male_only: str
    student_friendly: str
    student_score: int
    shared_confidence: str
    possible_house: str
    raw_relative_url: str
    raw_thumb_url: str
    scraped_at: str

    def normalized(self) -> NormalizedListing:
        return NormalizedListing(
            source=self.source,
            url=self.url,
            title=self.title,
            price_eur=self.price_eur,
            postcode=self.postcode,
            latitude=None,
            longitude=None,
            student_occupants="",
            worker_occupants="",
            scraped_at=self.scraped_at,
        )

    def explicit_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "source_listing_id": self.source_listing_id,
            "url": self.url,
            "title": self.title,
            "price_eur": self.price_eur,
            "listing_type": self.listing_type,
            "lodging_type": self.lodging_type,
            "lodging_type_label": self.lodging_type_label,
            "surface_sqm": self.surface_sqm,
            "room_count_label": self.room_count_label,
            "current_housemates": self.current_housemates,
            "furnished": self.furnished,
            "availability": self.availability,
            "published_at": self.published_at,
            "published_label": self.published_label,
            "city": self.city,
            "street": self.street,
            "postcode": self.postcode,
            "postcode_source": self.postcode_source,
            "company_name": self.company_name,
            "description": self.description,
            "raw_thumb_url": self.raw_thumb_url,
            "scraped_at": self.scraped_at,
        }


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape La Carte des Colocs Lyon listings.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--json-output", default=str(DEFAULT_JSON_OUTPUT))
    parser.add_argument("--max-results", type=int, default=0, help="Limit exported rows. 0 means all captured results.")
    parser.add_argument("--max-batches", type=int, default=0, help="Limit captured API batches. 0 means keep fetching until the site runs out of results.")
    parser.add_argument("--wait-ms", type=int, default=WAIT_MS, help="How long to wait for the page's listing request.")
    return parser.parse_args()


def looks_shared(record: dict[str, Any]) -> bool:
    listing_type = str(record.get("listing_type") or "")
    lodging_type = str(record.get("lodging_type") or "")
    title = compact(str(record.get("main_title") or ""))
    description = compact(str(record.get("description_truncated") or ""))

    if listing_type in SHARED_TYPES:
        return True
    if record.get("housemates"):
        return True
    haystack = f"{title} {description}".lower()
    if "coloc" in haystack or "chambre" in haystack:
        return True
    if lodging_type == "house" and "maison" in haystack:
        return True
    return False


def fetch_json(url: str) -> dict[str, Any]:
    request = Request(url, headers=GEOCODER_HEADERS)
    with urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8", errors="replace"))


def arrondissement_to_postcode(text: str) -> str | None:
    normalized = text.lower().replace("ème", "e").replace("eme", "e")
    match = re.search(r"\blyon(?:\s+|-)(1er|[1-9]e?|[1-9])\b", normalized)
    if not match:
        return None
    raw = match.group(1)
    if raw == "1er":
        arrondissement = 1
    else:
        arrondissement = int(re.sub(r"\D", "", raw))
    if 1 <= arrondissement <= 9:
        return f"6900{arrondissement}"
    return None


def geocode_postcode(street: str, city: str, cache: dict[str, str | None]) -> str | None:
    query = compact(f"{street}, {city}")
    if not query:
        return None
    if query in cache:
        return cache[query]

    url = f"{GEOCODER_URL}?q={quote_plus(query)}&limit=1"
    try:
        payload = fetch_json(url)
    except Exception:
        cache[query] = None
        return None

    features = payload.get("features") or []
    if not features:
        cache[query] = None
        return None

    properties = features[0].get("properties") or {}
    geocoded_postcode = properties.get("postcode")
    geocoded_city = compact(str(properties.get("city") or ""))
    geocoded_street = compact(str(properties.get("name") or ""))

    if not isinstance(geocoded_postcode, str) or not re.fullmatch(r"69\d{3}", geocoded_postcode):
        cache[query] = None
        return None
    if city and geocoded_city and city.lower() not in geocoded_city.lower():
        cache[query] = None
        return None
    if street and geocoded_street and not any(token.lower() in geocoded_street.lower() for token in street.split()[:2]):
        cache[query] = None
        return None

    cache[query] = geocoded_postcode
    return geocoded_postcode


def infer_postcode(title: str, street: str, city: str, description: str, geocode_cache: dict[str, str | None]) -> tuple[str | None, str]:
    haystack = " ".join([title, street, city, description])
    match = re.search(r"\b(69\d{3})\b", haystack)
    if match:
        return match.group(1), "explicit_text"

    arrondissement_postcode = arrondissement_to_postcode(haystack)
    if arrondissement_postcode:
        return arrondissement_postcode, "lyon_arrondissement"

    geocoded_postcode = geocode_postcode(street, city, geocode_cache)
    if geocoded_postcode:
        return geocoded_postcode, "geocoded_street_city"

    return None, "missing"


def score_student_friendliness(description: str) -> int:
    description_low = description.lower()
    score = 0
    for pattern in STUDENT_PATTERNS:
        if re.search(pattern, description_low):
            score += 1
    return score


def bool_label(value: bool) -> str:
    return "Yes" if value else "No"


def shared_confidence(record: dict[str, Any], description: str) -> str:
    listing_type = str(record.get("listing_type") or "")
    housemates = record.get("housemates")
    haystack = description.lower()
    if listing_type in {"flatshare", "coliving"}:
        return "High"
    if housemates:
        return "High"
    if "coloc" in haystack or "chambre" in haystack:
        return "Medium"
    return "Low"


def parse_listing_record(record: dict[str, Any], scraped_at: str, geocode_cache: dict[str, str | None]) -> Listing | None:
    if not looks_shared(record):
        return None

    title = compact(str(record.get("main_title") or "Untitled"))
    description = compact(str(record.get("description_truncated") or ""))
    street = compact(str(record.get("address_street") or ""))
    city = compact(str(record.get("address_city") or ""))
    relative_url = str(record.get("relative_url") or "")
    listing_url = urljoin(BASE_URL, relative_url)
    listing_type = str(record.get("listing_type") or "")
    lodging_type = str(record.get("lodging_type") or "")
    source_id = str(record.get("id") or record.get("url_token") or relative_url)
    student_score = score_student_friendliness(description)
    postcode, postcode_source = infer_postcode(title, street, city, description, geocode_cache)

    female_only = any(re.search(pattern, description.lower()) for pattern in FEMALE_ONLY_PATTERNS)
    male_only = any(re.search(pattern, description.lower()) for pattern in MALE_ONLY_PATTERNS)

    return Listing(
        source="la_carte_des_colocs",
        source_listing_id=source_id,
        url=listing_url,
        title=title,
        price_eur=record.get("cost_total_rent"),
        listing_type=listing_type,
        lodging_type=lodging_type,
        lodging_type_label=compact(str(record.get("lodging_type_string") or "")),
        surface_sqm=record.get("lodging_surface"),
        room_count_label=compact(str(record.get("lodging_size_string") or "")),
        current_housemates=record.get("housemates"),
        furnished=record.get("furnished"),
        availability=compact(str(record.get("lodging_availability_string") or "")),
        published_at=str(record.get("published_at") or ""),
        published_label=compact(str(record.get("published_at_string") or "")),
        city=city,
        street=street,
        postcode=postcode,
        postcode_source=postcode_source,
        company_name=compact(str(record.get("company_name") or "")),
        description=description,
        female_only=bool_label(female_only),
        male_only=bool_label(male_only),
        student_friendly=bool_label(student_score > 0),
        student_score=student_score,
        shared_confidence=shared_confidence(record, description),
        possible_house=bool_label(lodging_type == "house" or "maison" in f"{title} {description}".lower()),
        raw_relative_url=relative_url,
        raw_thumb_url=str(record.get("thumb_url") or ""),
        scraped_at=scraped_at,
    )


def parse_results_payload(raw_text: str, scraped_at: str) -> tuple[list[Listing], dict[str, Any]]:
    payload = json.loads(raw_text)
    results_raw = payload.get("results")
    if isinstance(results_raw, str):
        results = json.loads(results_raw)
    elif isinstance(results_raw, list):
        results = results_raw
    else:
        results = []

    parsed: list[Listing] = []
    geocode_cache: dict[str, str | None] = {}
    for record in results:
        listing = parse_listing_record(record, scraped_at, geocode_cache)
        if listing is not None:
            parsed.append(listing)

    meta = {
        "results_count": payload.get("results_count"),
        "batch_size": payload.get("batch_size"),
        "captured_count": len(results),
        "kept_count": len(parsed),
    }
    return parsed, meta


def build_offset_patch_script(offset: int) -> str:
    return f"""
(() => {{
  const targetOffset = {offset};
  const patchBody = (body) => {{
    try {{
      if (!body || typeof body !== "string") return body;
      const obj = JSON.parse(body);
      if (obj && obj.listing_search && obj.listing_search.filters) {{
        obj.listing_search.filters.offset = targetOffset;
        return JSON.stringify(obj);
      }}
    }} catch (error) {{
    }}
    return body;
  }};

  const originalFetch = window.fetch;
  window.fetch = function(input, init) {{
    const url = typeof input === "string" ? input : input && input.url;
    if (url && url.includes("/listing_search/list_results") && init) {{
      init = Object.assign({{}}, init, {{ body: patchBody(init.body) }});
    }}
    return originalFetch.call(this, input, init);
  }};

  const originalOpen = XMLHttpRequest.prototype.open;
  const originalSend = XMLHttpRequest.prototype.send;
  XMLHttpRequest.prototype.open = function(method, url) {{
    this.__lcdc_url = url;
    return originalOpen.apply(this, arguments);
  }};
  XMLHttpRequest.prototype.send = function(body) {{
    if (this.__lcdc_url && this.__lcdc_url.includes("/listing_search/list_results")) {{
      body = patchBody(body);
    }}
    return originalSend.call(this, body);
  }};
}})();
"""


def decode_results(raw_text: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    payload = json.loads(raw_text)
    results_raw = payload.get("results")
    if isinstance(results_raw, str):
        results = json.loads(results_raw)
    elif isinstance(results_raw, list):
        results = results_raw
    else:
        results = []
    return results, payload


def capture_single_batch(browser: Any, offset: int, wait_ms: int, include_page_snapshot: bool) -> tuple[str, str, str]:
    context = browser.new_context(
        locale="fr-FR",
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
    )
    page = context.new_page()
    if offset:
        page.add_init_script(build_offset_patch_script(offset))

    try:
        with page.expect_response(
            lambda response: response.url.endswith("/listing_search/list_results"),
            timeout=max(15_000, wait_ms + 5_000),
        ) as response_info:
            page.goto(START_URL, wait_until="domcontentloaded", timeout=60_000)

        response = response_info.value
        page.wait_for_timeout(min(wait_ms, 1_500))
        page_title = page.title() if include_page_snapshot else ""
        page_html = page.content() if include_page_snapshot else ""
        raw_response = response.text()
    finally:
        context.close()

    if response.status != 200 or not raw_response.lstrip().startswith("{"):
        raise RuntimeError(f"Could not capture a valid listing JSON batch at offset {offset}.")

    return raw_response, page_title, page_html


def capture_live_batches(wait_ms: int, max_batches: int) -> tuple[list[dict[str, Any]], str, str]:
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Run `pip install playwright` and install Chromium.")

    page_title = ""
    page_html = ""
    batches: list[dict[str, Any]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        offset = 0
        batch_index = 0
        expected_total: int | None = None
        batch_size: int | None = None

        while True:
            raw_response, current_title, current_html = capture_single_batch(
                browser=browser,
                offset=offset,
                wait_ms=wait_ms,
                include_page_snapshot=batch_index == 0,
            )
            if batch_index == 0:
                page_title = current_title
                page_html = current_html

            results, payload = decode_results(raw_response)
            expected_total = int(payload.get("results_count") or expected_total or 0) or expected_total
            batch_size = int(payload.get("batch_size") or batch_size or len(results) or 30)
            batches.append(
                {
                    "offset": offset,
                    "results_count": len(results),
                    "raw_response": raw_response,
                }
            )

            batch_index += 1
            if max_batches > 0 and batch_index >= max_batches:
                break
            if not results or len(results) < batch_size:
                break

            next_offset = offset + batch_size
            if expected_total is not None and next_offset >= expected_total:
                break
            offset = next_offset

        browser.close()

    return batches, page_title, page_html


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def export_excel(path: Path, listings: list[Listing], meta: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listings"

    headers = [
        "Source",
        "Listing ID",
        "Title",
        "URL",
        "Price EUR",
        "Listing Type",
        "Lodging Type",
        "Surface sqm",
        "Room Count",
        "Current Housemates",
        "Furnished",
        "Availability",
        "Published",
        "City",
        "Street",
        "Postcode",
        "Postcode Source",
        "Company",
        "Description",
        "Thumbnail",
        "Scraped At",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for listing in listings:
        ws.append(
            [
                listing.source,
                listing.source_listing_id,
                listing.title,
                listing.url,
                listing.price_eur,
                listing.listing_type,
                listing.lodging_type_label or listing.lodging_type,
                listing.surface_sqm,
                listing.room_count_label,
                listing.current_housemates,
                "Yes" if listing.furnished else "No" if listing.furnished is False else "",
                listing.availability,
                listing.published_label or listing.published_at,
                listing.city,
                listing.street,
                listing.postcode,
                listing.postcode_source,
                listing.company_name,
                listing.description,
                listing.raw_thumb_url,
                listing.scraped_at,
            ]
        )

    widths = {
        "A": 18,
        "B": 12,
        "C": 28,
        "D": 44,
        "E": 11,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 12,
        "J": 18,
        "K": 11,
        "L": 18,
        "M": 18,
        "N": 12,
        "O": 28,
        "P": 10,
        "Q": 16,
        "R": 18,
        "S": 80,
        "T": 36,
        "U": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    meta_sheet = wb.create_sheet("Run Meta")
    meta_sheet.append(["Key", "Value"])
    for key, value in meta.items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        meta_sheet.append([key, value])
    meta_sheet.column_dimensions["A"].width = 24
    meta_sheet.column_dimensions["B"].width = 40

    wb.save(path)


def build_capture_payload(listings: list[Listing], meta: dict[str, Any], title: str) -> dict[str, Any]:
    normalized = [listing.normalized().to_dict() for listing in listings]
    return {
        "source": "la_carte_des_colocs",
        "start_url": START_URL,
        "page_title": title,
        "meta": meta,
        "listings": [listing.explicit_dict() for listing in listings],
        "normalized_listings": normalized,
    }


def main() -> int:
    args = parse_args()
    output_path = Path(args.output).expanduser().resolve()
    json_output_path = Path(args.json_output).expanduser().resolve()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_CAPTURE_DIR.mkdir(parents=True, exist_ok=True)

    scraped_at = utc_now_iso()
    batch_payloads, page_title, page_html = capture_live_batches(wait_ms=args.wait_ms, max_batches=args.max_batches)

    listings_by_id: dict[str, Listing] = {}
    batch_summaries: list[dict[str, Any]] = []
    total_available = 0
    batch_size = 0
    for batch_payload in batch_payloads:
        batch_listings, batch_meta = parse_results_payload(batch_payload["raw_response"], scraped_at=scraped_at)
        batch_summaries.append(
            {
                "offset": batch_payload["offset"],
                "captured_count": batch_meta["captured_count"],
                "kept_count": batch_meta["kept_count"],
            }
        )
        total_available = max(total_available, int(batch_meta.get("results_count") or 0))
        batch_size = max(batch_size, int(batch_meta.get("batch_size") or 0))
        for listing in batch_listings:
            listings_by_id.setdefault(listing.source_listing_id, listing)

    listings = list(listings_by_id.values())
    meta = {
        "results_count": total_available,
        "batch_size": batch_size,
        "batches_fetched": len(batch_payloads),
        "captured_count": sum(item["captured_count"] for item in batch_summaries),
        "kept_count": len(listings),
        "deduped_count": sum(item["kept_count"] for item in batch_summaries) - len(listings),
        "batch_summaries": batch_summaries,
    }
    listings.sort(
        key=lambda item: (
            -item.student_score,
            item.female_only == "Yes",
            item.price_eur if item.price_eur is not None else 10**9,
            item.title,
        )
    )

    if args.max_results > 0:
        listings = listings[: args.max_results]

    capture_payload = build_capture_payload(listings, meta, page_title)
    export_excel(output_path, listings, meta)
    write_json(json_output_path, capture_payload)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_dir = RAW_CAPTURE_DIR / timestamp
    raw_dir.mkdir(parents=True, exist_ok=True)
    for batch_payload in batch_payloads:
        offset = int(batch_payload["offset"])
        filename = f"listing_response_offset_{offset:04d}.json"
        (raw_dir / filename).write_text(batch_payload["raw_response"], encoding="utf-8")
    (raw_dir / "page.html").write_text(page_html, encoding="utf-8")
    (raw_dir / "page_title.txt").write_text(page_title, encoding="utf-8")

    print(f"Saved Excel output to: {output_path}")
    print(f"Saved JSON output to: {json_output_path}")
    print(f"Saved raw capture to: {raw_dir}")
    print(
        f"Captured {meta['captured_count']} live results across {meta['batches_fetched']} batches, "
        f"kept {meta['kept_count']} shared listings."
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        raise SystemExit(130)
