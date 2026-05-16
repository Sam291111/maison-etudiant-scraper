"""
ImmoJeune Lyon shared-accommodation scraper
==========================================

What it does:
- Searches both Lyon `colocation` and `location-etudiant`
- Uses the site filters for `Chambre` + `Meuble`
- Keeps shared listings and excludes studio-style results
- Enriches listings from the detail page with student signals
- Exports a ranked Excel file

Run:
  python3 immojeune_scraper.py
  python3 immojeune_scraper.py --max-pages 4

Dependencies:
  pip install beautifulsoup4 openpyxl
"""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from urllib.parse import urlencode, quote_plus
from urllib.request import Request, urlopen

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.immojeune.com"
CITY_SLUG = "lyon-69"
BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from shared.listing_schema import NormalizedListing

OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_FILE = str(OUTPUT_DIR / "immojeune_lyon_shared_housing.xlsx")
JSON_OUTPUT_FILE = str(OUTPUT_DIR / "immojeune_lyon_shared_housing.json")
DEFAULT_MAX_PAGES = 6
REQUEST_DELAY = 1.0
TIMEOUT_SECONDS = 30
GEOCODER_URL = "https://data.geopf.fr/geocodage/search"

SEARCH_PARAMS = {
    "priceMin": 0,
    "priceMax": 3000,
    "surfaceMin": 0,
    "surfaceMax": 100,
    "propertyType[0]": "Chambre",
    "furnished": 1,
    "around": 1,
}

SOURCE_PATHS = [
    ("colocation", "Colocation"),
    ("location-etudiant", "Location etudiant"),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}

GOOD_LOCATIONS = {"69003", "69007", "69008", "69100", "69200", "69500", "69000"}
NEIGHBOURHOOD_MAP = {
    "69001": "1er - Presqu'ile",
    "69002": "2eme - Presqu'ile",
    "69003": "3eme - Part-Dieu / Guillotiere",
    "69004": "4eme - Croix-Rousse",
    "69005": "5eme - Vieux Lyon",
    "69006": "6eme - Brotteaux",
    "69007": "7eme - Jean Mace / Gerland",
    "69008": "8eme - Monplaisir",
    "69009": "9eme - Vaise",
    "69100": "Villeurbanne",
    "69200": "Venissieux",
    "69500": "Bron",
}

STRONG_SHARED_BADGES = {"CHAMBRE", "COLOCATION"}
SOLO_BADGES = {"STUDIO", "T1", "T1 BIS", "T1 BIS DUPLEX"}
STUDENT_PATTERNS = [
    r"\betudiant",
    r"\betudiants",
    r"\bétudiant",
    r"\bétudiants",
    r"\bstudent",
    r"\bstudents",
    r"jeunes actifs",
    r"young professionals",
    r"proche de l[' ]universite",
    r"proche de l[' ]université",
    r"campus",
]
STREET_KEYWORDS = r"rue|avenue|cours|quai|route|boulevard|place|allee|allée|impasse|chemin"


@dataclass
class Listing:
    title: str
    url: str
    source: str
    badges: list[str]
    price_num: int | None
    price_str: str
    size_sqm: int | None
    size_str: str
    postcode: str
    neighbourhood: str
    near_campus: bool
    location_text: str
    summary_description: str
    address_text: str
    student_occupants: str
    worker_occupants: str
    all_occupants: str
    female_count: int
    male_count: int
    min_age: int | None
    max_age: int | None
    female_only_household: str
    student_count: int
    worker_count: int
    mixed_household: str
    student_score: int
    student_label: str
    availability: str
    charges: str
    deposit: str
    fees: str
    nearby_schools: str
    latitude: float | None
    longitude: float | None
    coordinate_source: str
    coordinate_confidence: str

    def explicit_dict(self) -> dict:
        return {
            "title": self.title,
            "url": self.url,
            "source": self.source,
            "badges": self.badges,
            "price_num": self.price_num,
            "price_str": self.price_str,
            "size_sqm": self.size_sqm,
            "size_str": self.size_str,
            "postcode": self.postcode,
            "neighbourhood": self.neighbourhood,
            "near_campus": self.near_campus,
            "location_text": self.location_text,
            "summary_description": self.summary_description,
            "address_text": self.address_text,
            "student_occupants": self.student_occupants,
            "worker_occupants": self.worker_occupants,
            "all_occupants": self.all_occupants,
            "female_count": self.female_count,
            "male_count": self.male_count,
            "min_age": self.min_age,
            "max_age": self.max_age,
            "female_only_household": self.female_only_household,
            "student_count": self.student_count,
            "worker_count": self.worker_count,
            "mixed_household": self.mixed_household,
            "student_score": self.student_score,
            "student_label": self.student_label,
            "availability": self.availability,
            "charges": self.charges,
            "deposit": self.deposit,
            "fees": self.fees,
            "nearby_schools": self.nearby_schools,
            "latitude": self.latitude,
            "longitude": self.longitude,
            "coordinate_source": self.coordinate_source,
            "coordinate_confidence": self.coordinate_confidence,
        }

    def normalized(self) -> NormalizedListing:
        return NormalizedListing(
            source="immojeune",
            url=self.url,
            title=self.title,
            price_eur=self.price_num,
            postcode=self.postcode or None,
            latitude=self.latitude,
            longitude=self.longitude,
            student_occupants=self.student_occupants if self.student_occupants != "Not listed" else "",
            worker_occupants=self.worker_occupants if self.worker_occupants != "Not listed" else "",
        )


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def fetch_html(url: str) -> str:
    request = Request(url, headers=HEADERS)
    with urlopen(request, timeout=TIMEOUT_SECONDS) as response:
        return response.read().decode("utf-8", errors="replace")


def fetch_json(url: str) -> dict:
    request = Request(url, headers=HEADERS)
    with urlopen(request, timeout=TIMEOUT_SECONDS) as response:
        return json.loads(response.read().decode("utf-8", errors="replace"))


def build_search_url(source_path: str, page_num: int) -> str:
    base_path = f"/{source_path}/{CITY_SLUG}.html" if page_num == 1 else f"/{source_path}/{CITY_SLUG}/{page_num}"
    return f"{BASE_URL}{base_path}?{urlencode(SEARCH_PARAMS)}"


def parse_badges(card: BeautifulSoup) -> list[str]:
    return [compact(node.get_text(" ", strip=True)) for node in card.select(".badge")]


def decode_listing_url(title_node: BeautifulSoup | None) -> str | None:
    if title_node is None:
        return None
    if title_node.name == "a":
        href = title_node.get("href")
        if href:
            return href if href.startswith("http") else f"{BASE_URL}{href}"
    encoded = title_node.get("data-encoded-link")
    if encoded:
        decoded = base64.b64decode(encoded).decode("utf-8")
        return decoded if decoded.startswith("http") else f"{BASE_URL}{decoded}"
    return None


def extract_price(text: str) -> tuple[int | None, str]:
    match = re.search(r"(\d[\d\s]*)\s*€", text)
    if not match:
        return None, "N/A"
    price_num = int(match.group(1).replace(" ", ""))
    return price_num, f"{price_num} EUR"


def extract_size(text: str) -> tuple[int | None, str]:
    match = re.search(r"(\d+)\s*m²", text)
    if not match:
        return None, "N/A"
    size_num = int(match.group(1))
    return size_num, f"{size_num} m²"


def extract_postcode(text: str) -> str:
    match = re.search(r"\b(69\d{3})\b", text)
    return match.group(1) if match else "69000"


def looks_shared(badges: Iterable[str], title: str, url: str, description: str) -> bool:
    badge_set = {badge.upper() for badge in badges}
    title_up = title.upper()
    desc_up = description.upper()
    url_up = url.upper()

    if badge_set & STRONG_SHARED_BADGES:
        return True
    if "/COLOCATION/" in url_up:
        return True
    if "CHAMBRE" in title_up or "COLOCATION" in title_up:
        return True
    if "COLOCATION" in desc_up or "COLOCATAIRES" in desc_up:
        return True
    return False


def looks_solo(badges: Iterable[str], title: str, description: str) -> bool:
    badge_set = {badge.upper() for badge in badges}
    title_up = title.upper()
    desc_up = description.upper()

    if badge_set & SOLO_BADGES:
        return True
    if "STUDIO" in title_up or "STUDIO" in desc_up:
        return True
    if re.search(r"\bT1\b|\bT1 BIS\b", title_up):
        return True
    return False


def parse_listing_cards(html: str, source_label: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    cards = soup.select("#resultsajax .card.col")
    parsed: list[dict] = []

    for card in cards:
        badges = parse_badges(card)
        title_node = card.select_one("p.title a, p.title span.obflink")
        title = compact(title_node.get_text(" ", strip=True)) if title_node else "Untitled"
        url = decode_listing_url(title_node)
        if not url:
            continue

        desc_node = card.select_one("p.description")
        description = compact(desc_node.get_text(" ", strip=True)) if desc_node else ""
        location_node = card.select_one(".geo")
        location_text = compact(location_node.get_text(" ", strip=True).replace("Ville", "")) if location_node else ""
        price_num, price_str = extract_price(card.get_text(" ", strip=True))
        size_num, size_str = extract_size(card.get_text(" ", strip=True))
        postcode = extract_postcode(location_text)

        if not looks_shared(badges, title, url, description):
            continue
        if looks_solo(badges, title, description):
            continue

        parsed.append(
            {
                "title": title,
                "url": url,
                "source": source_label,
                "badges": badges,
                "price_num": price_num,
                "price_str": price_str,
                "size_num": size_num,
                "size_str": size_str,
                "postcode": postcode,
                "location_text": location_text,
                "summary_description": description,
            }
        )

    return parsed


def count_people(description: str, role: str) -> int:
    return len(re.findall(rf"\b(?:Male|Female),\s*{role}\b", description, re.I))


def parse_occupants(description: str) -> list[dict[str, int | str]]:
    occupants: list[dict[str, int | str]] = []
    pattern = re.compile(r"\b(?:\d+\)\s*)?(Male|Female),\s*(Student|Worker),\s*(\d{2})\b", re.I)
    for gender, role, age in pattern.findall(description):
        occupants.append(
            {
                "gender": gender.title(),
                "role": role.title(),
                "age": int(age),
            }
        )
    return occupants


def format_occupants(occupants: list[dict[str, int | str]], role: str | None = None) -> str:
    selected = [person for person in occupants if role is None or person["role"] == role]
    if not selected:
        return "Not listed"
    return " | ".join(f"{str(person['gender'])[0]}{person['age']}" for person in selected)


def extract_detail_fields(html: str, fallback: dict) -> Listing:
    soup = BeautifulSoup(html, "html.parser")
    title_node = soup.select_one("h1")
    title = compact(title_node.get_text(" ", strip=True)) if title_node else fallback["title"]

    desc_block = soup.select_one(".item.description")
    full_description = compact(desc_block.get_text(" ", strip=True)) if desc_block else fallback["summary_description"]

    availability = "Unknown"
    address_text = fallback["location_text"]
    for node in soup.select("aside h3"):
        value = compact(node.get_text(" ", strip=True))
        if "Disponible" in value:
            availability = value
            break

    title_meta = [compact(node.get_text(" ", strip=True)) for node in soup.select(".item.title p.flex.grey")]
    if title_meta:
        address_text = title_meta[0]

    charges = deposit = fees = ""
    sidebar = soup.select_one("#partial-advert-description")
    if sidebar:
        sidebar_text = compact(sidebar.get_text(" ", strip=True))
        charges_match = re.search(r"Charges\s*(.+?)\s*Frais", sidebar_text, re.I)
        fees_match = re.search(r"Frais de dossier [^A-Za-z0-9]*\s*(.+?)\s*D[eé]p[oô]t", sidebar_text, re.I)
        deposit_match = re.search(r"D[eé]p[oô]t de garantie [^A-Za-z0-9]*\s*(.+)$", sidebar_text, re.I)
        charges = compact(charges_match.group(1)) if charges_match else ""
        fees = compact(fees_match.group(1)) if fees_match else ""
        deposit = compact(deposit_match.group(1)) if deposit_match else ""

    school_names = [
        compact(node.get_text(" ", strip=True))
        for node in soup.select(".item.school li a")
    ]
    nearby_schools = " | ".join(school_names[:4])

    occupants = parse_occupants(full_description)
    student_count = sum(1 for person in occupants if person["role"] == "Student")
    worker_count = sum(1 for person in occupants if person["role"] == "Worker")
    female_count = sum(1 for person in occupants if person["gender"] == "Female")
    male_count = sum(1 for person in occupants if person["gender"] == "Male")
    ages = [int(person["age"]) for person in occupants]
    min_age = min(ages) if ages else None
    max_age = max(ages) if ages else None
    female_only_household = "Yes" if occupants and female_count == len(occupants) else "No"

    student_score = 0
    text_for_scoring = f"{title} {full_description}"
    for pattern in STUDENT_PATTERNS:
        if re.search(pattern, text_for_scoring, re.I):
            student_score += 1
    if student_count:
        student_score += 2 + student_count
    if worker_count:
        student_score -= 1
    if nearby_schools:
        student_score += 1

    if student_count and not worker_count:
        mixed_household = f"{student_count} student(s), no workers listed"
    elif student_count or worker_count:
        mixed_household = f"{student_count} student(s), {worker_count} worker(s)"
    else:
        mixed_household = "No current-housemate data shown"

    if student_score >= 6:
        student_label = "Strong student fit"
    elif student_score >= 3:
        student_label = "Likely student-friendly"
    elif student_score >= 1:
        student_label = "Possible student fit"
    else:
        student_label = "Unknown"

    postcode = fallback["postcode"]
    neighbourhood = NEIGHBOURHOOD_MAP.get(postcode, postcode)

    return Listing(
        title=title,
        url=fallback["url"],
        source=fallback["source"],
        badges=fallback["badges"],
        price_num=fallback["price_num"],
        price_str=fallback["price_str"],
        size_sqm=fallback["size_num"],
        size_str=fallback["size_str"],
        postcode=postcode,
        neighbourhood=neighbourhood,
        near_campus=postcode in GOOD_LOCATIONS,
        location_text=fallback["location_text"],
        summary_description=full_description,
        address_text=address_text,
        student_occupants=format_occupants(occupants, "Student"),
        worker_occupants=format_occupants(occupants, "Worker"),
        all_occupants=format_occupants(occupants),
        female_count=female_count,
        male_count=male_count,
        min_age=min_age,
        max_age=max_age,
        female_only_household=female_only_household,
        student_count=student_count,
        worker_count=worker_count,
        mixed_household=mixed_household,
        student_score=student_score,
        student_label=student_label,
        availability=availability,
        charges=charges or "Unknown",
        deposit=deposit or "Unknown",
        fees=fees or "Unknown",
        nearby_schools=nearby_schools or "None listed",
        latitude=None,
        longitude=None,
        coordinate_source="Not geocoded",
        coordinate_confidence="None",
    )


def clean_address_text(value: str) -> str:
    return compact(value.replace(" - ", ", ").replace(" ,", ","))


def infer_street_from_title(title: str) -> str:
    patterns = [
        r"chambre dans le (.+)",
        r"chambre dans la (.+)",
        r"chambre dans l[' ](.+)",
        r"colocation - (.+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, title, re.I)
        if match:
            return compact(match.group(1))
    return ""


def has_numbered_street_address(text: str) -> bool:
    return bool(re.search(rf"\b\d+\s+(?:{STREET_KEYWORDS})\b", text, re.I))


def build_geocode_queries(listing: Listing) -> list[tuple[str, str, str]]:
    queries: list[tuple[str, str, str]] = []
    address = clean_address_text(listing.address_text)
    title_street = infer_street_from_title(listing.title)

    if has_numbered_street_address(address):
        queries.append((address, "Full address", "High"))

    if title_street:
        street_query = compact(f"{title_street} {listing.postcode} Lyon")
        queries.append((street_query, "Street + postcode", "Medium"))
    elif re.search(rf"\b(?:{STREET_KEYWORDS})\b", address, re.I):
        queries.append((address, "Street + postcode", "Medium"))

    unique_queries: list[tuple[str, str, str]] = []
    seen = set()
    for query in queries:
        if query[0] in seen:
            continue
        seen.add(query[0])
        unique_queries.append(query)
    return unique_queries


def geocode_listing(listing: Listing, cache: dict[str, tuple[float, float, str, str]]) -> Listing:
    queries = build_geocode_queries(listing)
    for query_text, source_label, confidence in queries:
        cached = cache.get(query_text)
        if cached:
            listing.latitude, listing.longitude, listing.coordinate_source, listing.coordinate_confidence = cached
            return listing

        url = f"{GEOCODER_URL}?q={quote_plus(query_text)}&limit=1"
        try:
            payload = fetch_json(url)
        except Exception:
            continue

        features = payload.get("features") or []
        if not features:
            continue

        coordinates = features[0].get("geometry", {}).get("coordinates")
        if not coordinates or len(coordinates) < 2:
            continue

        longitude, latitude = coordinates[0], coordinates[1]
        cache[query_text] = (latitude, longitude, source_label, confidence)
        listing.latitude = latitude
        listing.longitude = longitude
        listing.coordinate_source = source_label
        listing.coordinate_confidence = confidence
        return listing

    return listing


def scrape_source(source_path: str, source_label: str, max_pages: int) -> list[dict]:
    found: list[dict] = []
    seen_urls: set[str] = set()

    for page_num in range(1, max_pages + 1):
        url = build_search_url(source_path, page_num)
        print(f"  {source_label} page {page_num}: {url}")
        html = fetch_html(url)
        listings = parse_listing_cards(html, source_label)

        if not listings:
            print("    No shared listings found on this page, stopping this source.")
            break

        new_count = 0
        for listing in listings:
            if listing["url"] in seen_urls:
                continue
            seen_urls.add(listing["url"])
            found.append(listing)
            new_count += 1

        print(f"    {new_count} new shared listing(s)")
        time.sleep(REQUEST_DELAY)

    return found


def enrich_listings(card_listings: list[dict]) -> list[Listing]:
    enriched: list[Listing] = []
    seen_urls: set[str] = set()
    geocode_cache: dict[str, tuple[float, float, str, str]] = {}

    for index, listing in enumerate(card_listings, start=1):
        if listing["url"] in seen_urls:
            continue
        seen_urls.add(listing["url"])
        print(f"  Detail {index}/{len(card_listings)}: {listing['url']}")
        try:
            html = fetch_html(listing["url"])
            enriched_listing = extract_detail_fields(html, listing)
            enriched_listing = geocode_listing(enriched_listing, geocode_cache)
            enriched.append(enriched_listing)
        except Exception as exc:
            print(f"    Detail fetch failed: {exc}")
            neighbourhood = NEIGHBOURHOOD_MAP.get(listing["postcode"], listing["postcode"])
            enriched.append(
                Listing(
                    title=listing["title"],
                    url=listing["url"],
                    source=listing["source"],
                    badges=listing["badges"],
                    price_num=listing["price_num"],
                    price_str=listing["price_str"],
                    size_sqm=listing["size_num"],
                    size_str=listing["size_str"],
                    postcode=listing["postcode"],
                    neighbourhood=neighbourhood,
                    near_campus=listing["postcode"] in GOOD_LOCATIONS,
                    location_text=listing["location_text"],
                    summary_description=listing["summary_description"],
                    address_text=listing["location_text"],
                    student_occupants="Not listed",
                    worker_occupants="Not listed",
                    all_occupants="Not listed",
                    female_count=0,
                    male_count=0,
                    min_age=None,
                    max_age=None,
                    female_only_household="No",
                    student_count=0,
                    worker_count=0,
                    mixed_household="Detail fetch failed",
                    student_score=0,
                    student_label="Unknown",
                    availability="Unknown",
                    charges="Unknown",
                    deposit="Unknown",
                    fees="Unknown",
                    nearby_schools="Unknown",
                    latitude=None,
                    longitude=None,
                    coordinate_source="Not geocoded",
                    coordinate_confidence="None",
                )
            )
        time.sleep(REQUEST_DELAY)

    return enriched


def sort_key(listing: Listing) -> tuple:
    return (
        -(listing.student_score),
        listing.near_campus is False,
        listing.price_num if listing.price_num is not None else 999999,
        listing.postcode,
        listing.title.lower(),
    )


def build_excel(listings: list[Listing], output_file: str) -> None:
    Path(output_file).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shared Housing"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    strong_fill = PatternFill("solid", start_color="E2EFDA")
    mixed_fill = PatternFill("solid", start_color="FFF2CC")
    campus_fill = PatternFill("solid", start_color="DDEBF7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="CCCCCC")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    columns = [
        "#",
        "Title",
        "Source",
        "Badges",
        "Price",
        "Size",
        "Postcode",
        "Neighbourhood",
        "Near Campus?",
        "Student Fit",
        "Current Mix",
        "Students (G/A)",
        "Workers (G/A)",
        "All Occupants",
        "Female Only?",
        "Age Range",
        "Available",
        "Charges",
        "Deposit",
        "Fees",
        "Nearby Schools",
        "Latitude",
        "Longitude",
        "Coord Source",
        "Coord Confidence",
        "Link",
    ]
    widths = [4, 38, 16, 26, 10, 9, 10, 25, 14, 18, 26, 18, 18, 18, 24, 12, 12, 18, 12, 12, 12, 34, 12, 12, 18, 16, 14]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws["A1"] = f"ImmoJeune Lyon shared housing | scraped {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    strong_count = sum(1 for listing in listings if listing.student_label == "Strong student fit")
    ws["A2"] = (
        f"{len(listings)} listings | {strong_count} strong student-fit listings | "
        "Shared-room focused, studio-style listings excluded"
    )
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = center

    for col_num, (header, width) in enumerate(zip(columns, widths), start=1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = width

    ws.freeze_panes = "A4"

    for row_num, listing in enumerate(listings, start=4):
        values = [
            row_num - 3,
            listing.title,
            listing.source,
            " | ".join(listing.badges),
            listing.price_str,
            listing.size_str,
            listing.postcode,
            listing.neighbourhood,
            "Yes" if listing.near_campus else "Further out",
            listing.student_label,
            listing.mixed_household,
            listing.student_occupants,
            listing.worker_occupants,
            listing.all_occupants,
            listing.female_only_household,
            f"{listing.min_age}-{listing.max_age}" if listing.min_age is not None and listing.max_age is not None else "Not listed",
            listing.availability,
            listing.charges,
            listing.deposit,
            listing.fees,
            listing.nearby_schools,
            listing.latitude,
            listing.longitude,
            listing.coordinate_source,
            listing.coordinate_confidence,
            listing.url,
        ]

        if listing.student_label == "Strong student fit":
            fill = strong_fill
        elif listing.near_campus:
            fill = campus_fill
        elif listing.student_count and listing.worker_count:
            fill = mixed_fill
        else:
            fill = None

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = left if col_num in {2, 4, 8, 10, 11, 12, 13, 14, 15, 18, 19, 20, 24, 25} else center
            if fill:
                cell.fill = fill
            if col_num == 26 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "View"
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

    legend_row = len(listings) + 6
    ws.cell(row=legend_row, column=1, value="Legend").font = Font(name="Arial", bold=True, size=9)
    legend_rows = [
        ("Green", "Strong student fit based on housemate mix and description", strong_fill),
        ("Blue", "Near Lyon campus areas", campus_fill),
        ("Yellow", "Mixed student/worker household", mixed_fill),
    ]
    for offset, (label, desc, fill) in enumerate(legend_rows, start=1):
        cell_label = ws.cell(row=legend_row + offset, column=1, value=label)
        cell_desc = ws.cell(row=legend_row + offset, column=2, value=desc)
        cell_label.font = cell_desc.font = Font(name="Arial", size=9)
        cell_label.fill = fill

    wb.save(output_file)


def scrape_all(max_pages: int) -> list[Listing]:
    merged_by_url: dict[str, dict] = {}

    print("Collecting filtered search pages...")
    for source_path, source_label in SOURCE_PATHS:
        try:
            source_cards = scrape_source(source_path, source_label, max_pages=max_pages)
        except Exception as exc:
            print(f"  Failed to scrape {source_label}: {exc}")
            continue
        for card in source_cards:
            existing = merged_by_url.get(card["url"])
            if existing:
                existing_sources = {part.strip() for part in existing["source"].split(" | ")}
                existing_sources.add(card["source"])
                existing["source"] = " | ".join(sorted(existing_sources))
                continue
            merged_by_url[card["url"]] = card

    merged_cards = list(merged_by_url.values())
    print(f"\nFound {len(merged_cards)} unique shared listing card(s)")
    enriched = enrich_listings(merged_cards)
    enriched.sort(key=sort_key)
    return enriched


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape shared housing in Lyon from ImmoJeune.")
    parser.add_argument("--max-pages", type=int, default=DEFAULT_MAX_PAGES, help="Pages per source to scan.")
    parser.add_argument("--output", default=OUTPUT_FILE, help="Excel output filename.")
    parser.add_argument("--json-output", default=JSON_OUTPUT_FILE, help="JSON output filename.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    print("Scraping ImmoJeune Lyon shared housing...")
    listings = scrape_all(max_pages=args.max_pages)
    if not listings:
        raise SystemExit("No shared listings found.")
    build_excel(listings, args.output)
    payload = {
        "source": "immojeune",
        "start_url": f"{BASE_URL}/colocation/{CITY_SLUG}.html",
        "meta": {
            "max_pages": args.max_pages,
            "listing_count": len(listings),
            "strong_student_fit_count": sum(1 for item in listings if item.student_label == "Strong student fit"),
        },
        "listings": [item.explicit_dict() for item in listings],
        "normalized_listings": [item.normalized().to_dict() for item in listings],
    }
    Path(args.json_output).expanduser().resolve().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nSaved {len(listings)} listing(s) to {args.output}")
    print(f"Saved JSON output to {args.json_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
