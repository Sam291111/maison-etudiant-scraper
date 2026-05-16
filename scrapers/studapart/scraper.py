"""
Studapart Lyon shared-accommodation scraper
==========================================

What it does:
- opens the live Lyon Studapart page in a browser session
- captures the site's own first search API request
- replays that request across pages
- keeps only listings that are explicitly shared-compatible
- exports conservative Excel + JSON outputs with only source-confirmed fields

Run:
  python3 scrapers/studapart/scraper.py
  python3 scrapers/studapart/scraper.py --max-pages 2

Dependencies:
  pip install beautifulsoup4 openpyxl playwright
  python3 -m playwright install
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
except ImportError:
    PlaywrightTimeoutError = None
    sync_playwright = None


BASE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BASE_DIR.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from shared.listing_schema import NormalizedListing, utc_now_iso
from scrapers.studapart.recon import DEFAULT_URL, HEADERS  # type: ignore[import-not-found]


BASE_URL = "https://www.studapart.com"
OUTPUT_DIR = BASE_DIR / "outputs"
DEFAULT_OUTPUT = OUTPUT_DIR / "studapart_lyon_shared_listings.xlsx"
DEFAULT_JSON_OUTPUT = OUTPUT_DIR / "studapart_lyon_shared_listings.json"
RAW_CAPTURE_DIR = OUTPUT_DIR / "raw_capture"


@dataclass
class Listing:
    source: str
    source_listing_id: str
    raw_id: str | None
    url: str
    title: str
    ad_type: str
    announcement_type: str
    residence_announcement_type: str
    price_eur: int | None
    full_accommodation_price_eur: int | None
    address: str
    full_address: str
    city: str
    postcode: str | None
    latitude: float | None
    longitude: float | None
    property_type: str
    surface_sqm: int | None
    rooms_count: int | None
    bedrooms_count: int | None
    bathroom_count: int | None
    max_tenants: int | None
    total_beds: int | None
    rented_by_room: bool | None
    all_rooms_available: bool | None
    fully_rentable: bool | None
    coliving: bool | None
    furnished: bool | None
    housing_assistance: bool | None
    private_bathroom: bool | None
    private_kitchen: bool | None
    without_visit: bool | None
    adapted_for_reduced_mobility: bool | None
    lessor_type: str
    owner_name: str
    residence_name: str
    tenant_status: str
    available_from: str
    available_until: str
    min_stay_months: int | None
    max_stay_months: int | None
    flexibility_days: int | None
    description: str
    canonical_url_fr: str
    canonical_url_en: str
    raw_media_count: int | None
    scraped_at: str

    def explicit_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "source_listing_id": self.source_listing_id,
            "raw_id": self.raw_id,
            "url": self.url,
            "title": self.title,
            "ad_type": self.ad_type,
            "announcement_type": self.announcement_type,
            "residence_announcement_type": self.residence_announcement_type,
            "price_eur": self.price_eur,
            "full_accommodation_price_eur": self.full_accommodation_price_eur,
            "address": self.address,
            "full_address": self.full_address,
            "city": self.city,
            "postcode": self.postcode,
            "latitude": self.latitude,
            "longitude": self.longitude,
            "property_type": self.property_type,
            "surface_sqm": self.surface_sqm,
            "rooms_count": self.rooms_count,
            "bedrooms_count": self.bedrooms_count,
            "bathroom_count": self.bathroom_count,
            "max_tenants": self.max_tenants,
            "total_beds": self.total_beds,
            "rented_by_room": self.rented_by_room,
            "all_rooms_available": self.all_rooms_available,
            "fully_rentable": self.fully_rentable,
            "coliving": self.coliving,
            "furnished": self.furnished,
            "housing_assistance": self.housing_assistance,
            "private_bathroom": self.private_bathroom,
            "private_kitchen": self.private_kitchen,
            "without_visit": self.without_visit,
            "adapted_for_reduced_mobility": self.adapted_for_reduced_mobility,
            "lessor_type": self.lessor_type,
            "owner_name": self.owner_name,
            "residence_name": self.residence_name,
            "tenant_status": self.tenant_status,
            "available_from": self.available_from,
            "available_until": self.available_until,
            "min_stay_months": self.min_stay_months,
            "max_stay_months": self.max_stay_months,
            "flexibility_days": self.flexibility_days,
            "description": self.description,
            "canonical_url_fr": self.canonical_url_fr,
            "canonical_url_en": self.canonical_url_en,
            "raw_media_count": self.raw_media_count,
            "scraped_at": self.scraped_at,
        }

    def normalized(self) -> NormalizedListing:
        return NormalizedListing(
            source=self.source,
            url=self.url,
            title=self.title,
            price_eur=self.price_eur,
            postcode=self.postcode,
            latitude=self.latitude,
            longitude=self.longitude,
            student_occupants="",
            worker_occupants="",
            scraped_at=self.scraped_at,
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape explicitly shared-compatible Studapart listings in Lyon.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--json-output", default=str(DEFAULT_JSON_OUTPUT))
    parser.add_argument("--max-pages", type=int, default=0)
    parser.add_argument("--max-results", type=int, default=0)
    parser.add_argument("--wait-ms", type=int, default=5000)
    return parser.parse_args()


def bool_label(value: bool | None) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return ""


def epoch_to_iso_day(value: Any) -> str:
    if not isinstance(value, int):
        return ""
    return datetime.fromtimestamp(value, tz=timezone.utc).date().isoformat()


def first_availability(item: dict[str, Any]) -> dict[str, Any] | None:
    availabilities = item.get("availabilities")
    if not isinstance(availabilities, list) or not availabilities:
        return None
    valid = [entry for entry in availabilities if isinstance(entry, dict)]
    if not valid:
        return None
    valid.sort(key=lambda entry: (entry.get("start") or 0, entry.get("end") or 0))
    return valid[0]


def explicit_shared_candidate(item: dict[str, Any]) -> bool:
    announcement_type = str(item.get("announcementType") or "")
    if announcement_type in {"flat_share", "coliving", "homestay"}:
        return True
    if item.get("rentedByRoom") is True:
        return True
    max_tenants = item.get("maxTenantNumber")
    if isinstance(max_tenants, int) and max_tenants > 1:
        return True
    return False


def shared_sort_key(item: dict[str, Any]) -> tuple:
    announcement_type = str(item.get("announcementType") or "")
    explicit_type_rank = {
        "flat_share": 0,
        "coliving": 1,
        "homestay": 2,
        "rental": 3,
        "service": 4,
    }.get(announcement_type, 9)
    rented_by_room_rank = 0 if item.get("rentedByRoom") is True else 1
    house_rank = 0 if item.get("propertyType") == "house" else 1
    max_tenants = item.get("maxTenantNumber")
    max_tenants_rank = -(max_tenants or 0)
    price_rank = item.get("rentWithExpensesAmount")
    price_value = price_rank if isinstance(price_rank, int) else 10**9
    title = title_for_item(item)
    return (explicit_type_rank, rented_by_room_rank, house_rank, max_tenants_rank, price_value, title)


def title_for_item(item: dict[str, Any]) -> str:
    title = item.get("title")
    if isinstance(title, str) and title.strip():
        return title.strip()
    residence_name = item.get("residenceName")
    if isinstance(residence_name, str) and residence_name.strip():
        return residence_name.strip()
    canonical = item.get("canonicalUrls") or {}
    if isinstance(canonical, dict):
        fr_url = canonical.get("fr")
        if isinstance(fr_url, str) and fr_url:
            slug = fr_url.strip("/").split("/")[-2] if len(fr_url.strip("/").split("/")) >= 2 else ""
            if slug:
                return slug.replace("-", " ").strip()
    return "Untitled listing"


def capture_search_template(start_url: str, wait_ms: int) -> tuple[str, dict[str, Any], dict[str, Any]]:
    if sync_playwright is None:
        raise RuntimeError("Playwright is not installed. Run: python3 -m pip install playwright && python3 -m playwright install")

    captured_request: dict[str, Any] | None = None
    captured_response: dict[str, Any] | None = None

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(
            locale="fr-FR",
            user_agent=HEADERS["User-Agent"],
            viewport={"width": 1440, "height": 1100},
        )
        page = context.new_page()

        def on_response(response: Any) -> None:
            nonlocal captured_request, captured_response
            if response.url != "https://search-api.studapart.com/property":
                return
            content_type = response.headers.get("content-type") or ""
            if "application/json" not in content_type:
                return
            try:
                request_data = json.loads(response.request.post_data or "{}")
                response_data = json.loads(response.text())
            except Exception:
                return
            captured_request = request_data
            captured_response = response_data

        page.on("response", on_response)
        page.goto(start_url, wait_until="domcontentloaded", timeout=60_000)
        try:
            page.wait_for_load_state("networkidle", timeout=20_000)
        except PlaywrightTimeoutError:
            pass
        page.wait_for_timeout(wait_ms)
        html = page.content()
        browser.close()

    if captured_request is None or captured_response is None:
        raise RuntimeError("Could not capture Studapart property search request from the live page.")

    return html, captured_request, captured_response


def post_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    headers = dict(HEADERS)
    headers["Content-Type"] = "application/json"
    request = Request(url, data=body, headers=headers, method="POST")
    with urlopen(request, timeout=60) as response:
        raw = response.read().decode("utf-8")
    return json.loads(raw)


def fetch_all_pages(template_payload: dict[str, Any], first_response: dict[str, Any], max_pages: int = 0) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    property_url = "https://search-api.studapart.com/property"
    first_results = list(first_response.get("results") or [])
    page_size = len(first_results) or 1
    nb_hits = int(first_response.get("nbHits") or 0)
    total_pages = max(1, math.ceil(nb_hits / page_size)) if nb_hits and page_size else 1
    if max_pages > 0:
        total_pages = min(total_pages, max_pages)

    all_results = list(first_results)
    page_summaries = [
        {
            "page": 1,
            "result_count": len(first_results),
            "is_last_page": bool(first_response.get("isLastPage")),
        }
    ]

    seen_ids = {
        str(item.get("distinctId") or item.get("_id"))
        for item in first_results
        if item.get("distinctId") or item.get("_id")
    }

    current_page = 2
    while current_page <= total_pages and not first_response.get("isLastPage"):
        payload = json.loads(json.dumps(template_payload))
        if not isinstance(payload.get("filters"), dict):
            break
        payload["filters"]["page"] = current_page
        response_data = post_json(property_url, payload)
        page_results = list(response_data.get("results") or [])
        new_count = 0
        for item in page_results:
            key = str(item.get("distinctId") or item.get("_id"))
            if key and key not in seen_ids:
                seen_ids.add(key)
                all_results.append(item)
                new_count += 1
        page_summaries.append(
            {
                "page": current_page,
                "result_count": len(page_results),
                "new_unique_results": new_count,
                "is_last_page": bool(response_data.get("isLastPage")),
            }
        )
        if response_data.get("isLastPage"):
            break
        if len(page_results) == 0 or new_count == 0:
            break
        current_page += 1

    meta = {
        "nb_hits": nb_hits,
        "page_size": page_size,
        "pages_fetched": len(page_summaries),
        "captured_results": len(all_results),
        "page_summaries": page_summaries,
    }
    return all_results, meta, page_summaries


def build_listing(item: dict[str, Any], scraped_at: str) -> Listing:
    availability = first_availability(item) or {}
    canonical = item.get("canonicalUrls") or {}
    latitude = None
    longitude = None
    geoloc = item.get("geoloc")
    if isinstance(geoloc, dict):
        lat = geoloc.get("lat")
        lon = geoloc.get("lon")
        latitude = float(lat) if isinstance(lat, (int, float)) else None
        longitude = float(lon) if isinstance(lon, (int, float)) else None

    canonical_fr = canonical.get("fr") if isinstance(canonical, dict) else ""
    canonical_en = canonical.get("en") if isinstance(canonical, dict) else ""
    url = urljoin(BASE_URL, canonical_fr or canonical_en or "")
    tenant_status = item.get("tenantStatus")
    tenant_status_text = ", ".join(tenant_status) if isinstance(tenant_status, list) else str(tenant_status or "")

    return Listing(
        source="studapart",
        source_listing_id=str(item.get("distinctId") or item.get("_id") or ""),
        raw_id=str(item.get("_id")) if item.get("_id") is not None else None,
        url=url,
        title=title_for_item(item),
        ad_type=str(item.get("ad_type") or ""),
        announcement_type=str(item.get("announcementType") or ""),
        residence_announcement_type=str(item.get("residenceAnnouncementType") or ""),
        price_eur=item.get("rentWithExpensesAmount") if isinstance(item.get("rentWithExpensesAmount"), int) else None,
        full_accommodation_price_eur=(
            item.get("rentWithExpensesAmountFullAccommodation")
            if isinstance(item.get("rentWithExpensesAmountFullAccommodation"), int)
            else None
        ),
        address=str(item.get("address") or ""),
        full_address=str(item.get("full_address") or ""),
        city=str(item.get("city") or ""),
        postcode=str(item.get("zipcode")) if item.get("zipcode") is not None else None,
        latitude=latitude,
        longitude=longitude,
        property_type=str(item.get("propertyType") or ""),
        surface_sqm=item.get("propertySurface") if isinstance(item.get("propertySurface"), int) else None,
        rooms_count=item.get("roomsCount") if isinstance(item.get("roomsCount"), int) else None,
        bedrooms_count=item.get("bedroomsCount") if isinstance(item.get("bedroomsCount"), int) else None,
        bathroom_count=item.get("bathroomCount") if isinstance(item.get("bathroomCount"), int) else None,
        max_tenants=item.get("maxTenantNumber") if isinstance(item.get("maxTenantNumber"), int) else None,
        total_beds=item.get("totalBeds") if isinstance(item.get("totalBeds"), int) else None,
        rented_by_room=item.get("rentedByRoom") if isinstance(item.get("rentedByRoom"), bool) else None,
        all_rooms_available=item.get("allRoomsAvailable") if isinstance(item.get("allRoomsAvailable"), bool) else None,
        fully_rentable=item.get("fullyRentable") if isinstance(item.get("fullyRentable"), bool) else None,
        coliving=item.get("coliving") if isinstance(item.get("coliving"), bool) else None,
        furnished=item.get("isFurnished") if isinstance(item.get("isFurnished"), bool) else None,
        housing_assistance=item.get("housingAssistance") if isinstance(item.get("housingAssistance"), bool) else None,
        private_bathroom=item.get("privateBathroom") if isinstance(item.get("privateBathroom"), bool) else None,
        private_kitchen=item.get("privateKitchen") if isinstance(item.get("privateKitchen"), bool) else None,
        without_visit=item.get("withoutVisit") if isinstance(item.get("withoutVisit"), bool) else None,
        adapted_for_reduced_mobility=(
            item.get("adaptedForReducedMobility") if isinstance(item.get("adaptedForReducedMobility"), bool) else None
        ),
        lessor_type=str(item.get("lessorType") or ""),
        owner_name=str(item.get("ownerName") or ""),
        residence_name=str(item.get("residenceName") or ""),
        tenant_status=tenant_status_text,
        available_from=epoch_to_iso_day(availability.get("start")),
        available_until=epoch_to_iso_day(availability.get("end")),
        min_stay_months=availability.get("min") if isinstance(availability.get("min"), int) else None,
        max_stay_months=availability.get("max") if isinstance(availability.get("max"), int) else None,
        flexibility_days=availability.get("flexibility") if isinstance(availability.get("flexibility"), int) else None,
        description=str(item.get("description") or ""),
        canonical_url_fr=str(canonical_fr or ""),
        canonical_url_en=str(canonical_en or ""),
        raw_media_count=item.get("picturesNumber") if isinstance(item.get("picturesNumber"), int) else None,
        scraped_at=scraped_at,
    )


def export_excel(path: Path, listings: list[Listing], meta: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shared Listings"

    headers = [
        "Source",
        "Listing ID",
        "Raw ID",
        "Title",
        "URL",
        "Ad Type",
        "Announcement Type",
        "Residence Type",
        "Price EUR",
        "Full Accommodation Price EUR",
        "Address",
        "Full Address",
        "City",
        "Postcode",
        "Latitude",
        "Longitude",
        "Property Type",
        "Surface sqm",
        "Rooms Count",
        "Bedrooms Count",
        "Bathrooms",
        "Max Tenants",
        "Total Beds",
        "Rented By Room",
        "All Rooms Available",
        "Fully Rentable",
        "Coliving",
        "Furnished",
        "Housing Assistance",
        "Private Bathroom",
        "Private Kitchen",
        "Without Visit",
        "Reduced Mobility",
        "Lessor Type",
        "Owner Name",
        "Residence Name",
        "Tenant Status",
        "Available From",
        "Available Until",
        "Min Stay (months)",
        "Max Stay (months)",
        "Flexibility Days",
        "Description",
        "Canonical FR",
        "Canonical EN",
        "Picture Count",
        "Scraped At",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for listing in listings:
        ws.append(
            [
                listing.source,
                listing.source_listing_id,
                listing.raw_id,
                listing.title,
                listing.url,
                listing.ad_type,
                listing.announcement_type,
                listing.residence_announcement_type,
                listing.price_eur,
                listing.full_accommodation_price_eur,
                listing.address,
                listing.full_address,
                listing.city,
                listing.postcode,
                listing.latitude,
                listing.longitude,
                listing.property_type,
                listing.surface_sqm,
                listing.rooms_count,
                listing.bedrooms_count,
                listing.bathroom_count,
                listing.max_tenants,
                listing.total_beds,
                bool_label(listing.rented_by_room),
                bool_label(listing.all_rooms_available),
                bool_label(listing.fully_rentable),
                bool_label(listing.coliving),
                bool_label(listing.furnished),
                bool_label(listing.housing_assistance),
                bool_label(listing.private_bathroom),
                bool_label(listing.private_kitchen),
                bool_label(listing.without_visit),
                bool_label(listing.adapted_for_reduced_mobility),
                listing.lessor_type,
                listing.owner_name,
                listing.residence_name,
                listing.tenant_status,
                listing.available_from,
                listing.available_until,
                listing.min_stay_months,
                listing.max_stay_months,
                listing.flexibility_days,
                listing.description,
                listing.canonical_url_fr,
                listing.canonical_url_en,
                listing.raw_media_count,
                listing.scraped_at,
            ]
        )

    widths = {
        "A": 14,
        "B": 40,
        "C": 28,
        "D": 38,
        "E": 58,
        "F": 12,
        "G": 18,
        "H": 18,
        "I": 10,
        "J": 18,
        "K": 28,
        "L": 34,
        "M": 24,
        "N": 10,
        "O": 12,
        "P": 12,
        "Q": 14,
        "R": 12,
        "S": 12,
        "T": 14,
        "U": 12,
        "V": 12,
        "W": 10,
        "X": 14,
        "Y": 16,
        "Z": 14,
        "AA": 10,
        "AB": 12,
        "AC": 16,
        "AD": 16,
        "AE": 14,
        "AF": 16,
        "AG": 18,
        "AH": 28,
        "AI": 18,
        "AJ": 18,
        "AK": 14,
        "AL": 14,
        "AM": 14,
        "AN": 90,
        "AO": 44,
        "AP": 44,
        "AQ": 12,
        "AR": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    meta_sheet = wb.create_sheet("Run Meta")
    meta_sheet.append(["Key", "Value"])
    for key, value in meta.items():
        meta_sheet.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value])
    meta_sheet.column_dimensions["A"].width = 24
    meta_sheet.column_dimensions["B"].width = 60

    wb.save(path)


def main() -> int:
    args = parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    RAW_CAPTURE_DIR.mkdir(parents=True, exist_ok=True)

    scraped_at = utc_now_iso()
    html, template_payload, first_response = capture_search_template(DEFAULT_URL, wait_ms=args.wait_ms)
    raw_results, page_meta, _page_summaries = fetch_all_pages(
        template_payload=template_payload,
        first_response=first_response,
        max_pages=args.max_pages,
    )

    kept_raw = [item for item in raw_results if explicit_shared_candidate(item)]
    kept_raw.sort(key=shared_sort_key)
    listings = [build_listing(item, scraped_at=scraped_at) for item in kept_raw]

    if args.max_results > 0:
        listings = listings[: args.max_results]

    output_path = Path(args.output).expanduser().resolve()
    json_output_path = Path(args.json_output).expanduser().resolve()

    meta = {
        **page_meta,
        "kept_count": len(listings),
        "excluded_count": len(raw_results) - len(kept_raw),
        "capture_url": DEFAULT_URL,
    }
    payload = {
        "source": "studapart",
        "start_url": DEFAULT_URL,
        "meta": meta,
        "listings": [item.explicit_dict() for item in listings],
        "normalized_listings": [item.normalized().to_dict() for item in listings],
    }

    export_excel(output_path, listings, meta)
    json_output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_dir = RAW_CAPTURE_DIR / timestamp
    raw_dir.mkdir(parents=True, exist_ok=True)
    (raw_dir / "page.html").write_text(html, encoding="utf-8")
    (raw_dir / "template_request.json").write_text(json.dumps(template_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (raw_dir / "first_response.json").write_text(json.dumps(first_response, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Saved Excel output to: {output_path}")
    print(f"Saved JSON output to: {json_output_path}")
    print(f"Saved raw capture to: {raw_dir}")
    print(f"Captured {meta['captured_results']} Studapart results and kept {meta['kept_count']} explicit shared listings.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
